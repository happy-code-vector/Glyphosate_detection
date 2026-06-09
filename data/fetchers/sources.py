"""
fetchers/cfia.py  /  fetchers/efsa.py  /  fetchers/fda.py

Structured government data sources — all Tier 2 (category aggregates).
All values computed from raw data files, nothing hardcoded.
"""

# ══════════════════════════════════════════════════════════════════════
# CFIA
# ══════════════════════════════════════════════════════════════════════

import io
import logging
import re
import zipfile
from pathlib import Path

import pandas as pd
import requests

from fetchers.base import BaseFetcher, download_file, SESSION, RAW_DATA_DIR, fetch_page
from db.database import normalize_category, build_dedup_key

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────
# CFIA — Canada Food Inspection Agency
# ─────────────────────────────────────────────────────────────────────

# The peer-reviewed scientific publication from CFIA contains structured
# supplementary tables (Tables S2–S7) with per-category breakdown.
# Primary source: pubs.acs.org/doi/10.1021/acs.jafc.9b07819
# The Open Government Portal CSV: open.canada.ca dataset
CFIA_GOV_PORTAL_URL = (
    "https://open.canada.ca/data/en/dataset/"
    "906cd35c-d396-4999-9a9f-f5351796661f"
)
# Direct CSV link (retrieved from portal — may need update if portal changes)
CFIA_CSV_URL = (
    "https://open.canada.ca/data/dataset/"
    "906cd35c-d396-4999-9a9f-f5351796661f/resource/"
    "glyphosate_food_residues_2015_2017.csv"
)
CFIA_FILENAME = "cfia_glyphosate_2015_2017.csv"
CFIA_SOURCE_URL = "https://inspection.canada.ca/en/food-safety-industry/food-chemistry-and-microbiology/food-safety-testing-reports-and-journal-articles/executive-summary"

CFIA_REPORTS = [
    {
        "label": "CFIA Glyphosate Testing 2015-2017",
        "type": "glyphosate_csv",
        "csv_url": (
            "https://open.canada.ca/data/dataset/"
            "906cd35c-d396-4999-9a9f-f5351796661f/resource/"
            "glyphosate_food_residues_2015_2017.csv"
        ),
        "portal_url": (
            "https://open.canada.ca/data/en/dataset/"
            "906cd35c-d396-4999-9a9f-f5351796661f"
        ),
        "filename": "cfia_glyphosate_2015_2017.csv",
        "published_date": "2019-04-01",
        "data_year": 2017,
    },
    {
        "label": "CFIA NCRMP 2016-2017",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/95a14ca0-706c-4422-ad42-b9e86998efbe",
        "filename": "cfia_ncrmp_2016_2017.csv",
        "published_date": "2018-01-01",
        "data_year": 2017,
    },
    {
        "label": "CFIA NCRMP 2017-2018",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/c87af563-b3f3-4048-96af-a5d39723ea6b",
        "filename": "cfia_ncrmp_2017_2018.csv",
        "published_date": "2019-01-01",
        "data_year": 2018,
    },
    {
        "label": "CFIA NCRMP 2018-2019",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/a2ea8989-2211-4d19-bc54-199dbd4c78ca",
        "filename": "cfia_ncrmp_2018_2019.csv",
        "published_date": "2020-01-01",
        "data_year": 2019,
    },
    {
        "label": "CFIA NCRMP 2019-2020",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/9e5211c8-c11f-4ebe-a7b2-65a6799a6032",
        "filename": "cfia_ncrmp_2019_2020.csv",
        "published_date": "2021-01-01",
        "data_year": 2020,
    },
    {
        "label": "CFIA NCRMP 2020-2021",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/a5cb7c3c-0371-4a20-ac9a-98fc4c3536bb",
        "filename": "cfia_ncrmp_2020_2021.csv",
        "published_date": "2022-01-01",
        "data_year": 2021,
    },
    {
        "label": "CFIA NCRMP 2021-2022",
        "type": "ncrmp_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/6567ac46-558e-4c95-ab93-e8326ddf8f90",
        "filename": "cfia_ncrmp_2021_2022.csv",
        "published_date": "2023-01-01",
        "data_year": 2022,
    },
    {
        "label": "CFIA Children's Food Project 2017",
        "type": "targeted_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/61a82716-e863-4c20-b1a7-c8e05e70e72d",
        "filename": "cfia_children_2017.csv",
        "published_date": "2018-01-01",
        "data_year": 2017,
    },
    {
        "label": "CFIA Selected Foods Survey 2018-2019",
        "type": "targeted_csv",
        "portal_url": "https://open.canada.ca/data/en/dataset/e4194282-102a-40ec-ac4c-0ce20e9a33cf",
        "filename": "cfia_selected_2018_2019.csv",
        "published_date": "2020-01-01",
        "data_year": 2019,
    },
]


class CFIAFetcher(BaseFetcher):
    SOURCE_NAME = "CFIA"

    def fetch(self) -> list[Path]:
        paths = []
        for report in CFIA_REPORTS:
            rtype = report["type"]
            if rtype == "glyphosate_csv":
                path = self._fetch_glyphosate_csv(report)
            else:
                path = self._fetch_portal_csv(report)
            if path:
                paths.append(path)
        return paths

    def _fetch_glyphosate_csv(self, report: dict) -> Path:
        """Download the original glyphosate-specific CSV."""
        try:
            return download_file(report["csv_url"], report["filename"])
        except Exception as e:
            logger.warning("Direct CSV URL failed: %s — trying portal page", e)
            return self._fetch_portal_csv(report)

    def _fetch_portal_csv(self, report: dict) -> Path | None:
        """Scrape Open Government Portal to find CSV download link."""
        from bs4 import BeautifulSoup

        cache_path = RAW_DATA_DIR / report["filename"]
        if cache_path.exists():
            logger.info("Cache hit: %s", report["filename"])
            return cache_path

        try:
            html = fetch_page(report["portal_url"])
            soup = BeautifulSoup(html, "html.parser")

            csv_links = [
                a["href"] for a in soup.find_all("a", href=True)
                if a["href"].endswith(".csv")
            ]
            if not csv_links:
                logger.warning("No CSV found on portal for %s", report["label"])
                return None

            csv_url = csv_links[0]
            if not csv_url.startswith("http"):
                csv_url = "https://open.canada.ca" + csv_url

            return download_file(csv_url, report["filename"])
        except Exception as e:
            logger.error("Failed to fetch %s: %s", report["label"], e)
            return None

    def parse(self, files: list[Path]) -> list[dict]:
        all_rows = []
        for path, report in zip(files, CFIA_REPORTS):
            if path is None:
                continue
            rtype = report["type"]
            if rtype == "glyphosate_csv":
                rows = self._parse_glyphosate_csv(path, report)
            elif rtype == "ncrmp_csv":
                rows = self._parse_ncrmp_csv(path, report)
            elif rtype == "targeted_csv":
                rows = self._parse_multi_pesticide_csv(path, report)
            else:
                logger.warning("Unknown CFIA report type: %s", rtype)
                continue
            all_rows.extend(rows)
        return all_rows

    def _parse_glyphosate_csv(self, csv_path: Path, report: dict) -> list[dict]:
        """Parse the original glyphosate-specific CSV (2015-2017).

        Extracts the sample year from the Date Sampled column so that
        2015, 2016, and 2017 data are stored under the correct data_year.
        """
        df = pd.read_csv(csv_path, low_memory=False)
        df.columns = [c.lower().strip().replace(" ", "_") for c in df.columns]
        logger.info("CFIA columns: %s", list(df.columns))

        product_col   = self._find_col(df, ["product", "produit", "commodity", "food_category"])
        component_col = self._find_col(df, ["component", "composant", "pesticide", "substance"])
        result_col    = self._find_col(df, ["result", "r_sultat", "concentration", "value"])
        if not result_col:
            result_col = next((c for c in df.columns if "result" in c or "rsultat" in c), None)
        unit_col = self._find_col(df, ["reportunit", "unit"])
        date_col = self._find_col(df, ["date_sampled", "date_d_chantillonage", "date"])
        if not date_col:
            # The bilingual column header has accents: date_sampled_¿_date_d'échantillonage
            date_col = next((c for c in df.columns if "date" in c.lower()), None)

        if not product_col or not component_col or not result_col:
            raise ValueError(
                f"Required columns not found in CFIA CSV. "
                f"Available: {list(df.columns)}."
            )

        gly_df = df[df[component_col].str.lower().str.contains("glyphosate", na=False)].copy()
        if gly_df.empty:
            raise ValueError("No glyphosate rows found in CFIA CSV")

        # Extract sample year from Date Sampled column
        if date_col and date_col in gly_df.columns:
            gly_df["_sample_year"] = pd.to_datetime(
                gly_df[date_col], errors="coerce"
            ).dt.year
        else:
            gly_df["_sample_year"] = report["data_year"]

        logger.info("CFIA: %d glyphosate sample rows", len(gly_df))

        conversion = 1.0
        original_unit = "µg/g"
        if unit_col:
            unit_val = str(gly_df[unit_col].iloc[0]).lower()
            if "mg/kg" in unit_val or "µg/g" in unit_val or "ug/g" in unit_val:
                conversion = 1000.0
                original_unit = unit_val

        rows = []
        for sample_year, year_group in gly_df.groupby("_sample_year"):
            sample_year = int(sample_year) if pd.notna(sample_year) else report["data_year"]

            product_stats = []
            for product, group in year_group.groupby(product_col):
                raw_cat = str(product).strip()
                if not raw_cat or raw_cat.lower() in ("nan", "total", "all"):
                    continue
                food_category = normalize_category(raw_cat)
                if not food_category:
                    continue
                values = pd.to_numeric(group[result_col], errors="coerce").fillna(0)
                product_stats.append({
                    "food_category": food_category,
                    "raw_cat": raw_cat,
                    "total": len(group),
                    "detected_values": values[values > 0].tolist(),
                })

            from collections import defaultdict
            by_category = defaultdict(lambda: {"total": 0, "detected": [], "raw_cats": []})
            for ps in product_stats:
                cat = ps["food_category"]
                by_category[cat]["total"] += ps["total"]
                by_category[cat]["detected"].extend(ps["detected_values"])
                by_category[cat]["raw_cats"].append(ps["raw_cat"])

            for food_category, stats in by_category.items():
                total = stats["total"]
                n_detected = len(stats["detected"])
                detection_rate = round(n_detected / total, 4) if total > 0 else None
                avg_ppb = round(sum(stats["detected"]) / n_detected * conversion, 2) if n_detected > 0 else None
                max_ppb = round(max(stats["detected"]) * conversion, 2) if stats["detected"] else None
                raw_cat = ", ".join(sorted(set(stats["raw_cats"])))

                rows.append({
                    "tier": 2,
                    "source_name": "CFIA",
                    "source_url": CFIA_SOURCE_URL,
                    "report_label": f"CFIA Glyphosate Testing {sample_year}",
                    "published_date": report["published_date"],
                    "data_year": sample_year,
                    "food_category": food_category,
                    "raw_category": raw_cat,
                    "samples_total": total,
                    "samples_detected": n_detected,
                    "detection_rate": detection_rate,
                    "avg_ppb": avg_ppb,
                    "max_ppb": max_ppb,
                    "original_unit": original_unit,
                    "unit_conversion": conversion,
                    "methodology_note": (
                        f"CFIA Glyphosate Testing {sample_year} (from 2015-2017 dataset). "
                        "Individual sample results aggregated by canonical food category. "
                        "LC-MS/MS method."
                    ),
                    "confidence": "medium",
                    "raw_file_path": str(csv_path),
                    "dedup_key": build_dedup_key("CFIA", food_category, sample_year),
                })

        logger.info("CFIA: parsed %d category rows from %s (split by year)", len(rows), report["label"])
        return rows

    def _parse_multi_pesticide_csv(self, csv_path: Path, report: dict) -> list[dict]:
        """
        Parse NCRMP or targeted survey CSVs that contain multiple pesticides.
        Filters for glyphosate rows only, then aggregates by food category.
        """
        try:
            df = pd.read_csv(csv_path, low_memory=False)
        except UnicodeDecodeError:
            df = pd.read_csv(csv_path, low_memory=False, encoding="latin-1")
        df.columns = [c.lower().strip().replace(" ", "_") for c in df.columns]
        logger.info("CFIA %s columns: %s", report["type"], list(df.columns))

        pest_col = self._find_col(df, [
            "pesticide", "substance", "param_name", "analyte",
            "chemical", "compound", "pesticide_name", "active_substance",
            "component", "composant",
        ])
        if not pest_col:
            logger.warning("CFIA: no pesticide column found in %s — skipping", report["label"])
            return []

        gly_df = df[df[pest_col].str.lower().str.contains("glyphosate", na=False)].copy()
        if gly_df.empty:
            logger.info("CFIA: no glyphosate rows in %s", report["label"])
            return []

        # Exclude AMPA metabolite
        gly_df = gly_df[~gly_df[pest_col].str.lower().str.contains("ampa", na=False)]
        if gly_df.empty:
            logger.info("CFIA: no glyphosate rows (only AMPA) in %s", report["label"])
            return []

        logger.info("CFIA: %d glyphosate rows in %s", len(gly_df), report["label"])

        product_col = self._find_col(df, [
            "product", "commodity", "food", "matrix", "product_name",
            "commodity_name", "food_product", "sample_type",
        ])
        if not product_col:
            logger.warning("CFIA: no product column found in %s — skipping", report["label"])
            return []

        result_col = self._find_col(df, [
            "result", "value", "concentration", "level", "residue",
            "detected_concentration", "measured_value", "amount",
        ])
        if not result_col:
            result_col = next((c for c in df.columns if "result" in c or "value" in c), None)

        unit_col = self._find_col(df, ["unit", "units", "result_unit", "report_unit"])

        conversion = 1000.0
        original_unit = "mg/kg"
        if unit_col:
            unit_val = str(gly_df[unit_col].iloc[0]).lower()
            if "ppb" in unit_val or "µg/kg" in unit_val or "ug/kg" in unit_val:
                conversion = 1.0
                original_unit = unit_val
            elif "ppm" in unit_val or "mg/kg" in unit_val:
                conversion = 1000.0
                original_unit = unit_val

        from collections import defaultdict
        by_category = defaultdict(lambda: {"total": 0, "detected": [], "raw_cats": []})

        for product, group in gly_df.groupby(product_col):
            raw_cat = str(product).strip()
            if not raw_cat or raw_cat.lower() in ("nan", "total", "all"):
                continue
            food_category = normalize_category(raw_cat)
            if not food_category:
                continue

            total = len(group)
            if result_col:
                values = pd.to_numeric(group[result_col], errors="coerce").fillna(0)
                detected = values[values > 0].tolist()
            else:
                detected = []

            by_category[food_category]["total"] += total
            by_category[food_category]["detected"].extend(detected)
            by_category[food_category]["raw_cats"].append(raw_cat)

        rows = []
        for food_category, stats in by_category.items():
            total = stats["total"]
            n_detected = len(stats["detected"])
            detection_rate = round(n_detected / total, 4) if total > 0 else None
            avg_ppb = round(sum(stats["detected"]) / n_detected * conversion, 2) if n_detected > 0 else None
            max_ppb = round(max(stats["detected"]) * conversion, 2) if stats["detected"] else None
            raw_cat = ", ".join(sorted(set(stats["raw_cats"])))

            rows.append({
                "tier": 2,
                "source_name": "CFIA",
                "source_url": report["portal_url"],
                "report_label": report["label"],
                "published_date": report["published_date"],
                "data_year": report["data_year"],
                "food_category": food_category,
                "raw_category": raw_cat,
                "samples_total": total,
                "samples_detected": n_detected,
                "detection_rate": detection_rate,
                "avg_ppb": avg_ppb,
                "max_ppb": max_ppb,
                "original_unit": original_unit,
                "unit_conversion": conversion,
                "methodology_note": (
                    f"{report['label']}. Multi-pesticide dataset filtered for glyphosate. "
                    "Individual sample results aggregated by canonical food category."
                ),
                "confidence": "medium",
                "raw_file_path": str(csv_path),
                "dedup_key": build_dedup_key("CFIA", food_category, report["data_year"]),
            })

        logger.info("CFIA: parsed %d category rows from %s", len(rows), report["label"])
        return rows

    def _parse_ncrmp_csv(self, csv_path: Path, report: dict) -> list[dict]:
        """
        Parse NCRMP summary CSV. These are pre-aggregated tables with
        unnamed columns. Food category is in a section header row ABOVE
        each data block — we look backwards from each glyphosate row to
        find the nearest commodity header.
        """
        try:
            df = pd.read_csv(csv_path, low_memory=False, skiprows=5)
        except UnicodeDecodeError:
            df = pd.read_csv(csv_path, low_memory=False, skiprows=5, encoding="latin-1")
        df.columns = [f"col_{i}" for i in range(len(df.columns))]
        logger.info("CFIA NCRMP rows: %d, columns: %d", len(df), len(df.columns))

        skip_labels = {"PESTICIDES", "METAL", "METALS", "VETERINARY DRUGS", "nan", ""}

        # Filter for glyphosate group + specific test
        gly_mask = (
            df["col_2"].astype(str).str.upper().str.strip() == "GLYPHOSATE"
        ) & (
            df["col_3"].astype(str).str.strip().str.lower() == "glyphosate"
        )
        gly_indices = df[gly_mask].index.tolist()
        if not gly_indices:
            logger.info("CFIA NCRMP: no glyphosate rows in %s", report["label"])
            return []

        logger.info("CFIA NCRMP: %d glyphosate rows in %s", len(gly_indices), report["label"])

        rows = []
        for gi in gly_indices:
            # Look backwards for nearest section header (food category)
            raw_cat = None
            for j in range(gi - 1, max(0, gi - 50), -1):
                col0 = str(df.iloc[j]["col_0"]).strip()
                col1 = df.iloc[j].get("col_1")
                if col0 and col0 not in skip_labels and (pd.isna(col1) or str(col1).strip() == "nan"):
                    raw_cat = col0
                    break

            if not raw_cat:
                continue

            food_category = normalize_category(raw_cat)
            if not food_category:
                logger.debug("CFIA NCRMP: no canonical category for '%s'", raw_cat)
                continue

            row = df.iloc[gi]
            total = int(pd.to_numeric(row.get("col_4"), errors="coerce") or 0)
            detected_val = pd.to_numeric(row.get("col_5"), errors="coerce")
            detected = int(detected_val) if pd.notna(detected_val) else 0
            avg_ppm = pd.to_numeric(row.get("col_6"), errors="coerce")
            max_ppm = pd.to_numeric(row.get("col_8"), errors="coerce")

            avg_ppb = round(float(avg_ppm) * 1000, 2) if pd.notna(avg_ppm) and avg_ppm > 0 else None
            max_ppb = round(float(max_ppm) * 1000, 2) if pd.notna(max_ppm) and max_ppm > 0 else None
            detection_rate = round(detected / total, 4) if total > 0 else None

            rows.append({
                "tier": 2,
                "source_name": "CFIA",
                "source_url": report["portal_url"],
                "report_label": report["label"],
                "published_date": report["published_date"],
                "data_year": report["data_year"],
                "food_category": food_category,
                "raw_category": raw_cat,
                "samples_total": total,
                "samples_detected": detected,
                "detection_rate": detection_rate,
                "avg_ppb": avg_ppb,
                "max_ppb": max_ppb,
                "original_unit": "ppm",
                "unit_conversion": 1000.0,
                "methodology_note": (
                    f"{report['label']}. NCRMP pre-aggregated summary data. "
                    f"Glyphosate-specific results for {raw_cat}. "
                    "Values in ppm converted to ppb."
                ),
                "confidence": "high",
                "raw_file_path": str(csv_path),
                "dedup_key": build_dedup_key("CFIA", food_category, raw_cat, report["data_year"]),
            })

        logger.info("CFIA NCRMP: parsed %d category rows from %s", len(rows), report["label"])
        return rows

    def _find_col(self, df, candidates: list[str]) -> str | None:
        for col in candidates:
            if col in df.columns:
                return col
        return None


# ══════════════════════════════════════════════════════════════════════
# EFSA
# ══════════════════════════════════════════════════════════════════════

# EFSA publishes enforcement data as XLSX on Zenodo.
# Table 2.2 contains individual MRL exceedance records with food matrix,
# substance name, and measured concentration (mg/kg).
EFSA_REPORTS = [
    # NOTE: EFSA 2016-2019 visualisation data is sample-level aggregate counts
    # without individual pesticide residue values — unusable for glyphosate ppb.
    # Only enforcement annexes (2020+) contain substance-specific MRL exceedance data.
    {
        "label": "EFSA EU Pesticide Residue Monitoring 2020",
        "zenodo_record": "6410774",
        "filename": "efsa_2020_enforcement.xlsx",
        "published_date": "2022-03-01",
        "data_year": 2020,
        "source_url": "https://zenodo.org/records/6410774",
        "format": "enforcement",
    },
    {
        "label": "EFSA EU Pesticide Residue Monitoring 2021",
        "zenodo_record": "7767236",
        "filename": "efsa_2021_enforcement.xlsx",
        "published_date": "2023-04-01",
        "data_year": 2021,
        "source_url": "https://zenodo.org/records/7767236",
        "format": "enforcement",
    },
    {
        "label": "EFSA EU Pesticide Residue Monitoring 2022",
        "zenodo_record": "10853986",
        "filename": "efsa_2022_enforcement.xlsx",
        "published_date": "2024-04-01",
        "data_year": 2022,
        "source_url": "https://zenodo.org/records/10853986",
        "format": "enforcement",
    },
    {
        "label": "EFSA EU Pesticide Residue Monitoring 2023",
        "zenodo_record": "14765085",
        "filename": "efsa_2023_enforcement.xlsx",
        "published_date": "2025-01-01",
        "data_year": 2023,
        "source_url": "https://zenodo.org/records/14765085",
        "format": "enforcement",
    },
    {
        "label": "EFSA EU Pesticide Residue Monitoring 2024",
        "zenodo_record": "18327007",
        "filename": "efsa_2024_enforcement.xlsx",
        "published_date": "2026-05-01",
        "data_year": 2024,
        "source_url": "https://zenodo.org/records/18327007",
        "format": "enforcement",
    },
]

EFSA_ZENODO_API = "https://zenodo.org/api/records/{record_id}"


class EFSAFetcher(BaseFetcher):
    SOURCE_NAME = "EFSA"

    def fetch(self) -> list[Path]:
        paths = []
        for report in EFSA_REPORTS:
            try:
                path = self._fetch_enforcement(report)
                paths.append(path)
            except Exception as e:
                logger.error(
                    "EFSA: failed to fetch %s: %s — skipping", report["label"], e
                )
        return paths

    def _fetch_enforcement(self, report: dict) -> Path:
        """Download enforcement or visualisation data XLSX from Zenodo."""
        cache_path = Path(__file__).parent.parent / "raw_data" / report["filename"]

        if cache_path.exists():
            logger.info("Cache hit: %s", cache_path.name)
            return cache_path

        record_id = report["zenodo_record"]
        api_url = EFSA_ZENODO_API.format(record_id=record_id)
        resp = SESSION.get(api_url, timeout=30)
        resp.raise_for_status()
        record_data = resp.json()

        files = record_data.get("files", [])
        fmt = report.get("format", "enforcement")

        if fmt == "enforcement":
            xlsx_files = [
                f for f in files
                if f.get("key", "").lower().endswith(".xlsx")
                and "enforcement" in f.get("key", "").lower()
            ]
            if not xlsx_files:
                xlsx_files = [f for f in files if f.get("key", "").lower().endswith(".xlsx")]
        else:
            xlsx_files = [
                f for f in files
                if f.get("key", "").lower().endswith(".xlsx")
                and "monitoring" in f.get("key", "").lower()
            ]
            if not xlsx_files:
                xlsx_files = [f for f in files if f.get("key", "").lower().endswith(".xlsx")]

        if not xlsx_files:
            raise RuntimeError(f"No XLSX files found in Zenodo record {record_id}")

        if fmt == "enforcement":
            data_file = min(xlsx_files, key=lambda f: f.get("size", 0))
        else:
            data_file = max(xlsx_files, key=lambda f: f.get("size", 0))

        download_url = data_file.get("links", {}).get("self")
        if not download_url:
            raise RuntimeError(f"No download URL for {data_file.get('key','')}")

        return download_file(download_url, report["filename"])

    def parse(self, files: list[Path]) -> list[dict]:
        all_rows = []
        file_map = {f.name: f for f in files}
        for report in EFSA_REPORTS:
            path = file_map.get(report["filename"])
            if path is None:
                continue
            fmt = report.get("format", "enforcement")
            if fmt == "visualisation":
                rows = self._parse_visualisation(path, report)
            else:
                rows = self._parse_enforcement(path, report)
            all_rows.extend(rows)
        return all_rows

    def _parse_enforcement(self, xlsx_path: Path, report: dict) -> list[dict]:
        """
        Parse EFSA enforcement data for ALL pesticide MRL exceedances.
        Sheet name varies by year: Table 2.2 (2023+), Table 3.2 (2020-2022).
        """
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheet_candidates = ["Table 2.2", "Table 3.2"]
        sheet_name = next((s for s in sheet_candidates if s in wb.sheetnames), None)
        wb.close()
        if not sheet_name:
            logger.warning(
                "EFSA: no enforcement sheet found in %s (sheets: %s)",
                xlsx_path.name, wb.sheetnames,
            )
            return []

        df = pd.read_excel(
            xlsx_path, sheet_name=sheet_name,
            header=None,
        )
        # Find the header row dynamically — look for "Substance Name"
        header_idx = None
        for idx, row in df.iterrows():
            vals = [str(v).strip() for v in row]
            if any("Substance Name" in v for v in vals):
                header_idx = idx
                break
        if header_idx is None:
            logger.warning("EFSA: no header row found in %s", xlsx_path.name)
            return []

        header = df.iloc[header_idx].tolist()
        df = df.iloc[header_idx + 1:].reset_index(drop=True)
        df.columns = [str(h).strip() for h in header]
        logger.info("EFSA %s columns: %s", sheet_name, list(df.columns))

        # Resolve column name variations across years
        substance_col = self._find_col(df, ["Substance Name", "Substance Name*"])
        result_col = self._find_col(df, ["Result (mg/kg)", "Result Value (mg/kg)"])
        if substance_col is None or result_col is None:
            logger.warning(
                "EFSA: missing expected columns in %s (substance=%s, result=%s)",
                xlsx_path.name, substance_col, result_col,
            )
            return []

        # Find the matrix column (food category)
        matrix_col = self._find_col(df, ["Food MATRIX Name", "Food Matrix Name", "Food MATRIX name"])
        if matrix_col is None:
            # Fallback: try substring match
            for col in df.columns:
                if "matrix" in str(col).lower() and "name" in str(col).lower():
                    matrix_col = col
                    break
        if matrix_col is None:
            logger.warning("EFSA: no food matrix column found in %s", xlsx_path.name)
            return []

        # Clean substance names
        df[substance_col] = df[substance_col].astype(str).str.strip()
        df = df[df[substance_col].str.lower() != 'nan']

        logger.info("EFSA: %d total rows in %s", len(df), xlsx_path.name)

        # Process ALL pesticides, grouped by (matrix, substance)
        rows = []
        for (matrix, substance), group in df.groupby([matrix_col, substance_col]):
            raw_cat = str(matrix).strip()
            pest_name = str(substance).strip().lower()
            if not raw_cat or raw_cat.lower() == 'nan':
                continue
            if not pest_name or pest_name == 'nan':
                continue

            food_category = normalize_category(raw_cat)
            if not food_category:
                continue

            total = len(group)
            values = pd.to_numeric(group[result_col], errors="coerce").fillna(0)
            detected = values[values > 0]
            n_detected = len(detected)

            # mg/kg → ppb (× 1000)
            avg_ppb = round(float(detected.mean()) * 1000, 2) if len(detected) > 0 else None
            max_ppb = round(float(detected.max()) * 1000, 2) if len(detected) > 0 else None

            rows.append({
                "tier": 2,
                "source_name": "EFSA",
                "source_url": report["source_url"],
                "report_label": report["label"],
                "published_date": report["published_date"],
                "data_year": report["data_year"],
                "food_category": food_category,
                "raw_category": raw_cat,
                "contaminant": pest_name,
                "samples_total": total,
                "samples_detected": n_detected,
                "detection_rate": round(n_detected / total, 4) if total > 0 else 0.0,
                "avg_ppb": avg_ppb,
                "max_ppb": max_ppb,
                "original_unit": "mg/kg",
                "unit_conversion": 1000.0,
                "methodology_note": (
                    "EFSA enforcement data: only samples exceeding MRL are reported. "
                    "Detection rate computed from exceedance samples only. "
                    f"{n_detected}/{total} MRL exceedance(s) for {pest_name} in {raw_cat} across EU member states."
                ),
                "confidence": "low",
                "raw_file_path": str(xlsx_path),
                "dedup_key": build_dedup_key("EFSA", food_category, pest_name, report["data_year"]),
            })

        logger.info("EFSA: parsed %d (category, pesticide) rows from %s", len(rows), xlsx_path.name)
        return rows

    def _parse_visualisation(self, xlsx_path: Path, report: dict) -> list[dict]:
        """
        Parse EFSA visualisation XLSX (2016-2017 format).
        Different structure than enforcement annexes — scans sheets for
        glyphosate rows with commodity and residue data.
        """
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        all_rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            headers = None
            for row in rows_iter:
                row_str = [str(c).lower() if c else "" for c in row]
                if any("glyphosate" in s for s in row_str):
                    headers = [str(c).strip() if c else "" for c in row]
                    break
                if (any("substance" in s for s in row_str)
                        and any(t in " ".join(row_str) for t in ["commodity", "product", "matrix"])):
                    headers = [str(c).strip() if c else "" for c in row]
                    break

            if not headers:
                continue

            headers_lower = [h.lower() for h in headers]

            substance_col = next((i for i, h in enumerate(headers_lower) if "substance" in h), None)
            commodity_col = next((i for i, h in enumerate(headers_lower) if any(t in h for t in ["commodity", "matrix", "product", "food"])), None)
            result_col = next((i for i, h in enumerate(headers_lower) if any(t in h for t in ["result", "value", "concentration", "mean", "level"])), None)
            samples_col = next((i for i, h in enumerate(headers_lower) if any(t in h for t in ["sample", "number", "total", "analysed"])), None)

            if substance_col is None or commodity_col is None:
                continue

            for row in rows_iter:
                if row[commodity_col] is None:
                    continue
                substance = str(row[substance_col] or "").lower()
                if "glyphosate" not in substance:
                    continue
                # Exclude AMPA (metabolite)
                words = substance.split()
                if any("ampa" in w for w in words):
                    continue

                raw_cat = str(row[commodity_col]).strip()
                food_category = normalize_category(raw_cat)
                if not food_category:
                    continue

                total = int(row[samples_col]) if samples_col is not None and row[samples_col] else None
                result_val = None
                if result_col is not None and row[result_col] is not None:
                    try:
                        result_val = float(row[result_col])
                    except (ValueError, TypeError):
                        pass

                avg_ppb = round(result_val * 1000, 2) if result_val else None
                max_ppb = avg_ppb

                all_rows.append({
                    "tier": 2,
                    "source_name": "EFSA",
                    "source_url": report["source_url"],
                    "report_label": report["label"],
                    "published_date": report["published_date"],
                    "data_year": report["data_year"],
                    "food_category": food_category,
                    "raw_category": raw_cat,
                    "samples_total": total or 1,
                    "samples_detected": 1,
                    "detection_rate": None,
                    "avg_ppb": avg_ppb,
                    "max_ppb": max_ppb,
                    "original_unit": "mg/kg",
                    "unit_conversion": 1000.0,
                    "methodology_note": (
                        f"EFSA visualisation data ({report['data_year']}). "
                        "Aggregated statistics per commodity. "
                        "Detection rate not available in this format."
                    ),
                    "confidence": "low",
                    "raw_file_path": str(xlsx_path),
                    "dedup_key": build_dedup_key("EFSA", food_category, report["data_year"]),
                })

        wb.close()
        logger.info("EFSA visualisation: parsed %d rows from %s", len(all_rows), xlsx_path.name)
        return all_rows

    def _find_col(self, df, candidates):
        for col in candidates:
            if col in df.columns:
                return col
        return None


# ══════════════════════════════════════════════════════════════════════
# FDA
# ══════════════════════════════════════════════════════════════════════

# FDA Pesticide Residue Monitoring Program — FY2023 uses aggregated
# CountryProductResidueData file with per-product per-chemical stats.
FDA_REPORTS = [
    {
        "label": "FDA Pesticide Monitoring FY2014",
        "year": 2014,
        "data_zip": "https://www.fda.gov/media/103551/download?attachment",
        "data_file": "CountryProductResidue2014.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2014",
        "published_date": "2017-03-01",
        "data_year": 2014,
    },
    {
        "label": "FDA Pesticide Monitoring FY2015",
        "year": 2015,
        "data_zip": "https://www.fda.gov/media/109108/download?attachment",
        "data_file": "CountryProductResidue2015.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2015",
        "published_date": "2017-11-01",
        "data_year": 2015,
    },
    {
        "label": "FDA Pesticide Monitoring FY2016",
        "year": 2016,
        "data_zip": "https://www.fda.gov/media/117091/download?attachment",
        "data_file": "CountryProductResidue2016.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2016",
        "published_date": "2018-10-01",
        "data_year": 2016,
    },
    {
        "label": "FDA Pesticide Monitoring FY2017",
        "year": 2017,
        "data_zip": "https://www.fda.gov/media/130342/download?attachment",
        "data_file": "CountryProductResidue2017.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2017",
        "published_date": "2019-09-01",
        "data_year": 2017,
    },
    {
        "label": "FDA Pesticide Monitoring FY2018",
        "year": 2018,
        "data_zip": "https://www.fda.gov/media/140798/download?attachment",
        "data_file": "CountryProductResidue2018.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2018",
        "published_date": "2020-09-01",
        "data_year": 2018,
    },
    {
        "label": "FDA Pesticide Monitoring FY2019",
        "year": 2019,
        "data_zip": "https://www.fda.gov/media/153148/download?attachment",
        "data_file": "CountryProductResidue2019.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2019",
        "published_date": "2021-10-01",
        "data_year": 2019,
    },
    {
        "label": "FDA Pesticide Monitoring FY2020",
        "year": 2020,
        "data_zip": "https://www.fda.gov/media/160408/download?attachment",
        "data_file": "CountryProductResidue2020.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2020",
        "published_date": "2022-08-01",
        "data_year": 2020,
    },
    {
        "label": "FDA Pesticide Monitoring FY2021",
        "year": 2021,
        "data_zip": "https://www.fda.gov/media/173178/download?attachment",
        "data_file": "CountryProductResidue2021.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2021",
        "published_date": "2023-10-01",
        "data_year": 2021,
    },
    {
        "label": "FDA Pesticide Monitoring FY2022",
        "year": 2022,
        "data_zip": "https://www.fda.gov/media/181378/download?attachment",
        "data_file": "CountryProductResidue2022.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2022",
        "published_date": "2024-09-01",
        "data_year": 2022,
    },
    {
        "label": "FDA Pesticide Monitoring FY2023",
        "year": 2023,
        "data_zip": "https://www.fda.gov/media/190132/download?attachment",
        "data_file": "CountryProductResidueData2023.txt",
        "sample_data_file": "SampleData2023.txt",
        "source_url": "https://www.fda.gov/food/pesticides/pesticide-residue-monitoring-report-and-data-fy-2023",
        "published_date": "2025-12-01",
        "data_year": 2023,
    },
]


class FDAFetcher(BaseFetcher):
    SOURCE_NAME = "FDA"

    def fetch(self) -> list[Path]:
        paths = []
        for report in FDA_REPORTS:
            year = report["year"]
            try:
                zip_path = download_file(report["data_zip"], f"fda_{year}_residue.zip")
            except Exception as e:
                logger.warning(
                    "FDA FY%s direct download failed (%s) — trying source page",
                    year, e
                )
                zip_path = self._fetch_via_source_page(report)
                if zip_path is None:
                    logger.error(
                        "FDA FY%s: could not download data — skipping this year", year
                    )
                    continue

            txt_path = Path(__file__).parent.parent / "raw_data" / report["data_file"]
            if not txt_path.exists():
                try:
                    with zipfile.ZipFile(zip_path) as zf:
                        names = zf.namelist()
                        year_str = str(year)
                        match = next(
                            (f for f in names
                             if "countryproductresidue" in f.lower()
                             and (year_str in f or f.lower() == "countryproductresiduedata.txt")),
                            None
                        )
                        if not match:
                            logger.error(
                                "CountryProductResidue file not found in FY%s zip. Available: %s — skipping year",
                                year, names
                            )
                            continue
                        # Save to the canonical filename expected by parser
                        data = zf.read(match)
                        txt_path.write_bytes(data)
                        logger.info("Extracted %s from %s", match, zip_path.name)
                except zipfile.BadZipFile:
                    logger.error("FDA FY%s: downloaded file is not a valid ZIP — skipping", year)
                    if zip_path.exists():
                        zip_path.unlink()
                    continue
            else:
                logger.info("Cache hit: %s", txt_path.name)
            paths.append(txt_path)
        return paths

    def _fetch_via_source_page(self, report: dict) -> Path | None:
        """Try to find the download link on the FDA source page."""
        try:
            from bs4 import BeautifulSoup
            html = fetch_page(report["source_url"])
            soup = BeautifulSoup(html, "html.parser")
            # Look for links to ZIP or media download files
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if ".zip" in href.lower() or "/media/" in href.lower():
                    if "download" in href.lower() or "sample" in href.lower() or "residue" in href.lower():
                        if not href.startswith("http"):
                            href = "https://www.fda.gov" + href
                        return download_file(href, f"fda_{report['year']}_residue.zip")
            logger.warning("No download link found on %s", report["source_url"])
            return None
        except Exception as e:
            logger.warning("Failed to scrape FDA source page for FY%s: %s", report["year"], e)
            return None

    def parse(self, files: list[Path]) -> list[dict]:
        all_rows = []
        # Match files to reports by data_file name
        file_map = {f.name: f for f in files}
        for report in FDA_REPORTS:
            path = file_map.get(report["data_file"])
            if path is None:
                continue
            rows = self._parse_fda(path, report)
            all_rows.extend(rows)

            # Tier 1: individual sample data (only 2023 has this)
            sample_file = report.get("sample_data_file")
            if sample_file:
                sample_path = RAW_DATA_DIR / sample_file
                if sample_path.exists():
                    t1_rows = self._parse_fda_samples(sample_path, report)
                    all_rows.extend(t1_rows)

        return all_rows

    def _parse_fda(self, data_path: Path, report: dict) -> list[dict]:
        """
        FDA CountryProductResidueData file is tab-delimited, pre-aggregated per
        product per chemical per country. Columns include:
          ProdName, ResName, Spls., Pos., Pos%, Mean, Minimum, Median, 90th, Maximum
        Values are in ppm (mg/kg) — convert to ppb (× 1000).
        """
        df = pd.read_csv(data_path, sep="\t", low_memory=False, encoding="latin-1")
        df.columns = [c.strip() for c in df.columns]
        logger.info("FDA columns: %s", list(df.columns))

        # Filter to GLYPHOSATE only (exclude N-ACETYLGLYPHOSATE and other metabolites)
        gly = df[
            df["ResName"].str.upper().str.strip() == "GLYPHOSATE"
        ].copy()

        if gly.empty:
            logger.warning("FDA: no glyphosate rows found")
            return []

        logger.info("FDA: %d glyphosate product-country rows", len(gly))

        # Aggregate across countries by product name → canonical category
        from collections import defaultdict
        by_category = defaultdict(lambda: {"total": 0, "detected": 0, "ppb_values": [], "raw_cats": []})

        for _, row in gly.iterrows():
            raw_cat = str(row["ProdName"]).strip()
            food_category = normalize_category(raw_cat)
            if not food_category:
                continue

            total = int(row.get("Spls.", 0) or 0)
            pos = int(row.get("Pos.", 0) or 0)
            mean_val = pd.to_numeric(row.get("Mean", 0), errors="coerce") or 0
            max_val = pd.to_numeric(row.get("Maximum", 0), errors="coerce") or 0

            cat = by_category[food_category]
            cat["total"] += total
            cat["detected"] += pos
            # FDA Mean is in ppm — convert to ppb
            if mean_val > 0:
                cat["ppb_values"].append(mean_val * 1000)
            if max_val > 0:
                cat["ppb_values"].append(max_val * 1000)
            cat["raw_cats"].append(raw_cat)

        rows = []
        for food_category, stats in by_category.items():
            if stats["total"] == 0:
                continue
            n_detected = stats["detected"]
            detection_rate = round(n_detected / stats["total"], 4)
            avg_ppb = round(sum(stats["ppb_values"]) / len(stats["ppb_values"]), 2) if stats["ppb_values"] else None
            max_ppb = round(max(stats["ppb_values"]), 2) if stats["ppb_values"] else None
            raw_cat = ", ".join(sorted(set(stats["raw_cats"])))

            rows.append({
                "tier": 2,
                "source_name": "FDA",
                "source_url": report["source_url"],
                "report_label": report["label"],
                "published_date": report["published_date"],
                "data_year": report["data_year"],
                "food_category": food_category,
                "raw_category": raw_cat,
                "samples_total": stats["total"],
                "samples_detected": n_detected,
                "detection_rate": detection_rate,
                "avg_ppb": avg_ppb,
                "max_ppb": max_ppb,
                "original_unit": "ppm",
                "unit_conversion": 1000.0,
                "methodology_note": (
                    "FDA Pesticide Residue Monitoring Program FY2023. "
                    "Aggregated across all countries of origin per product type. "
                    "Mean and Maximum from FDA summary stats (ppm converted to ppb)."
                ),
                "confidence": "high",
                "raw_file_path": str(data_path),
                "dedup_key": build_dedup_key("FDA", food_category, report["data_year"]),
            })

        logger.info("FDA: parsed %d category rows", len(rows))
        return rows

    def _parse_fda_samples(self, sample_path: Path, report: dict) -> list[dict]:
        """
        Parse FDA SampleData file for Tier 1 individual product test results.
        Each row is one sample with a specific product, residue finding, and unit.
        Columns: Year, SplNo, ProdName, Country, ResName, Found, Unit, Trace
        """
        df = pd.read_csv(sample_path, sep="\t", low_memory=False, encoding="latin-1")
        df.columns = [c.strip() for c in df.columns]

        gly = df[
            df["ResName"].str.upper().str.strip() == "GLYPHOSATE"
        ].copy()

        if gly.empty:
            logger.info("FDA samples: no glyphosate rows in %s", sample_path.name)
            return []

        logger.info("FDA samples: %d individual glyphosate samples", len(gly))

        rows = []
        for _, row in gly.iterrows():
            product_name = str(row.get("ProdName", "")).strip()
            if not product_name or product_name.lower() == "nan":
                continue

            food_category = normalize_category(product_name)
            if not food_category:
                continue

            found_val = pd.to_numeric(row.get("Found"), errors="coerce")
            is_trace = str(row.get("Trace", "")).strip().upper() == "T"
            country = str(row.get("Country", "")).strip()

            measured_ppb = round(found_val * 1000, 2) if pd.notna(found_val) and found_val > 0 else None
            below_detection = measured_ppb is None

            sample_id = str(row.get("SplNo", "")).strip()
            rows.append({
                "tier": 1,
                "source_name": "FDA",
                "source_url": report["source_url"],
                "report_label": report["label"],
                "published_date": report["published_date"],
                "data_year": report["data_year"],
                "food_category": food_category,
                "raw_category": product_name,
                "product_name": product_name,
                "measured_ppb": measured_ppb,
                "below_detection": below_detection,
                "country": country if country and country.lower() != "nan" else None,
                "methodology_note": (
                    f"FDA Pesticide Residue Monitoring Program FY{report['year']}. "
                    "Individual sample result. ppm converted to ppb."
                    + (" Trace detection (below LOQ)." if is_trace else "")
                ),
                "confidence": "high",
                "raw_file_path": str(sample_path),
                "dedup_key": build_dedup_key(
                    "FDA_T1", product_name, country, sample_id, report["data_year"]
                ),
            })

        logger.info("FDA samples: parsed %d Tier 1 product test rows", len(rows))
        return rows

    def _find_col(self, df, candidates):
        for col in candidates:
            if col in df.columns:
                return col
        return None
